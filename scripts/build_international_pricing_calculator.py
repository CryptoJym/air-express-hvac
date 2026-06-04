from __future__ import annotations

import csv
import json
import math
import re
import statistics
import subprocess
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path("/Users/jamesbrady/Projects/air-express-hvac")
DOWNLOADS = Path("/Users/jamesbrady/Downloads")
OUTPUT_DIR = ROOT / "output" / "spreadsheet"

TODAY = "2026-04-06"
DEFAULT_COMPUTE_SURCHARGE = 1.00
DEFAULT_TARGET_MARGIN = 0.35
DEFAULT_UNDERCUT = 0.01

GOODHIRE_URL = "https://docs.goodhire.com/pages/IntlTableNonUSCoverage3.html"
GBS_URL = (
    "https://www.globalbackgroundscreening.com/online-background-check/"
    "INTERNATIONAL-EMPLOYEE-SCREENING-Select-Country-For-Pricing-p303546142"
)
OWENS_COUNTRIES_URL = "https://www.owens.com/countries"
OWENS_CRIMINAL_ENDPOINTS = {
    14: "https://www.owens.com/api/gator-order/be/product?filter=productType.id=14&pageSize=200&page={page}",
    76: "https://www.owens.com/api/gator-order/be/product?filter=productType.id=76&pageSize=200&page={page}",
    51: "https://www.owens.com/api/gator-order/be/product?filter=productType.id=51&pageSize=200&page={page}",
}

INFORMDATA_SOURCE = DOWNLOADS / "InternationalPricingProposal_VuplicityInc.pdf"
INFORMDATA_BASE_CSV = DOWNLOADS / "InternationalPricing_market_floor_2026-03-31.csv"
OLD_OWENS_MAP_CSV = DOWNLOADS / "InternationalPricing_owens_comparison_2026-03-31.csv"
NEEYAMO_SOURCE = DOWNLOADS / "Neeyamo_iBGV_Ratecard_Vuplicity.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
CONTROL_FILL = PatternFill("solid", fgColor="DDEBF7")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
NEG_FILL = PatternFill("solid", fgColor="FDE9E7")
POS_FILL = PatternFill("solid", fgColor="E2F0D9")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


@dataclass
class PublicOffer:
    vendor: str
    price: float
    label: str
    source_url: str
    note: str = ""


def normalize_country(value: str | None) -> str:
    if not value:
        return ""
    normalized = (
        unicodedata.normalize("NFKD", str(value))
        .encode("ascii", "ignore")
        .decode("ascii")
        .upper()
    )
    normalized = normalized.replace("&", " AND ")
    normalized = normalized.replace("'", " ")
    normalized = re.sub(r"\bTHE\b", " ", normalized)
    normalized = re.sub(r"[^A-Z0-9]+", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()

    replacements = {
        "ANTIGUA BARBUDA": "ANTIGUA AND BARBUDA",
        "AZERBAJAN": "AZERBAIJAN",
        "BOSNIA HERZEGOVINA": "BOSNIA AND HERZEGOVINA",
        "BRISTISH VIRGIN ISLANDS": "BRITISH VIRGIN ISLANDS",
        "CABO VERDE CAPE VERDE": "CAPE VERDE",
        "COTE DIVOIRE": "COTE D IVOIRE",
        "CZECH REPUBLIC": "CZECHIA",
        "MACAO": "MACAU",
        "SWAZILAND": "ESWATINI",
        "U S VIRGIN ISLANDS": "VIRGIN ISLANDS US",
        "UNITED ARAB EMIRATES UAE": "UNITED ARAB EMIRATES",
        "VIRGIN ISLANDS U S": "VIRGIN ISLANDS US",
    }
    return replacements.get(normalized, normalized)


COUNTRY_ALIASES = {
    "ANTIGUA AND BARBUDA": ["ANTIGUA AND BARBUDA", "ANTIGUA BARBUDA", "ANTIGUA BARBUDA"],
    "BRITISH VIRGIN ISLANDS": ["BRITISH VIRGIN ISLANDS", "VIRGIN ISLANDS BRITISH"],
    "CAPE VERDE": ["CAPE VERDE", "CABO VERDE", "CABO VERDE CAPE VERDE"],
    "CONGO": ["CONGO", "CONGO BRAZZAVILLE", "REPUBLIC OF THE CONGO"],
    "COTE D IVOIRE": ["COTE D IVOIRE", "IVORY COAST"],
    "CYPRUS": ["CYPRUS", "CYPRUS REPUBLIC OF", "REPUBLIC OF CYPRUS"],
    "CZECHIA": ["CZECHIA", "CZECH REPUBLIC"],
    "ESWATINI": ["ESWATINI", "SWAZILAND"],
    "MACAU": ["MACAU", "MACAO"],
    "MICRONESIA": ["MICRONESIA", "CAROLINE ISLANDS", "FEDERATED STATES OF MICRONESIA"],
    "NORTH MACEDONIA": ["NORTH MACEDONIA", "MACEDONIA"],
    "PALESTINIAN TERRITORY": [
        "PALESTINIAN TERRITORY",
        "PALESTINIAN TERRITORIES",
        "PALESTINE",
    ],
    "SCOTLAND": ["SCOTLAND", "UNITED KINGDOM"],
    "UNITED ARAB EMIRATES": ["UNITED ARAB EMIRATES", "UNITED ARAB EMIRATES UAE", "UAE"],
    "VIRGIN ISLANDS US": ["VIRGIN ISLANDS US", "U S VIRGIN ISLANDS"],
}


def candidate_keys(country: str, preferred_label: str | None = None) -> list[str]:
    keys: list[str] = []
    if preferred_label:
        keys.append(normalize_country(preferred_label))
    base_key = normalize_country(country)
    keys.append(base_key)
    for alias in COUNTRY_ALIASES.get(base_key, []):
        keys.append(normalize_country(alias))
    deduped: list[str] = []
    for key in keys:
        if key and key not in deduped:
            deduped.append(key)
    return deduped


def parse_money(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\$?\s*(-?\d+(?:\.\d+)?)", str(value))
    return float(match.group(1)) if match else None


def safe_float(value: str | None) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def load_informdata_rows() -> list[dict]:
    rows: list[dict] = []
    with INFORMDATA_BASE_CSV.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            rows.append(
                {
                    "country": row["country"].strip(),
                    "scope": row["proposal_scope"].strip(),
                    "informdata_price": safe_float(row["direct_cost_usd"]),
                    "old_goodhire_label": row["goodhire_label"].strip(),
                    "old_gbs_label": row["gbs_label"].strip(),
                }
            )
    return rows


def load_old_owens_matches() -> dict[tuple[str, str], str]:
    mapping: dict[tuple[str, str], str] = {}
    with OLD_OWENS_MAP_CSV.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            mapping[(row["country"].strip(), "")] = row["owens_country_match"].strip()
    return mapping


def load_neeyamo_prices() -> dict[str, dict]:
    workbook = openpyxl.load_workbook(NEEYAMO_SOURCE, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    data: dict[str, dict] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        country = row[1]
        raw_price = row[3]
        price = parse_money(raw_price)
        if not country or price is None:
            continue
        key = normalize_country(str(country))
        entry = {
            "country": str(country).strip(),
            "price": price,
            "raw_price": str(raw_price).strip(),
        }
        # Keep the lowest listed criminal rate if there are duplicates.
        if key not in data or price < data[key]["price"]:
            data[key] = entry
    return data


def fetch_goodhire_offers() -> dict[str, PublicOffer]:
    response = requests.get(GOODHIRE_URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    offers: dict[str, PublicOffer] = {}
    for row in soup.select("tr"):
        cells = row.find_all("td")
        if len(cells) < 4:
            continue
        label = " ".join(cells[0].stripped_strings)
        price = parse_money(" ".join(cells[1].stripped_strings))
        if price is None:
            continue
        key = normalize_country(label)
        offers[key] = PublicOffer(
            vendor="GoodHire",
            price=price,
            label=label,
            source_url=GOODHIRE_URL,
            note="Current public international criminal table",
        )
    return offers


def fetch_gbs_offers() -> dict[str, list[PublicOffer]]:
    node_script = r"""
const { chromium } = require('playwright');
(async () => {
  const browser = await chromium.launch({ headless: true });
  const page = await browser.newPage({ viewport: { width: 1440, height: 900 } });
  await page.goto(
    'https://www.globalbackgroundscreening.com/online-background-check/INTERNATIONAL-EMPLOYEE-SCREENING-Select-Country-For-Pricing-p303546142',
    { waitUntil: 'domcontentloaded', timeout: 120000 }
  );
  await page.waitForTimeout(12000);
  const storeFrame = page.frames().find((frame) => frame.url().includes('/wix/app/store/'));
  const options = await storeFrame
    .locator('select')
    .nth(0)
    .locator('option')
    .evaluateAll((elements) =>
      elements.map((element) => ({
        text: (element.textContent || '').trim(),
        value: element.value,
      }))
    );
  console.log(JSON.stringify(options));
  await browser.close();
})().catch((error) => {
  console.error(error);
  process.exit(1);
});
"""
    raw_output = subprocess.check_output(["node", "-e", node_script], cwd=str(ROOT), text=True)
    options = json.loads(raw_output)

    offers: dict[str, list[PublicOffer]] = {}
    for option in options:
        text = option["text"].strip()
        if not text or text == "Please choose" or "SELECT COUNTRY" in text:
            continue
        price_match = re.search(r"\(\+\$(\d+(?:\.\d+)?)\)\s*$", text)
        if not price_match:
            continue
        price = float(price_match.group(1))
        label = re.sub(r"\s*\(\+\$\d+(?:\.\d+)?\)\s*$", "", text).strip()
        label = re.sub(r"\s+\([^()]*?(?:DAY|DAYS|YR|YRS|YEAR|YEARS)[^()]*\)\s*$", "", label).strip()
        label_key = normalize_country(label)
        base_key = re.sub(r"\bILLICIT ACTIVITY DATABASES ONLY\b", "", label_key).strip()
        base_key = re.sub(r"\bNATIONWIDE DBS CHECK BASIC\b", "", base_key).strip()
        base_key = re.sub(r"\bCRJMC LEVEL 2 NATIONWIDE\b", "", base_key).strip()
        base_key = re.sub(r"\b7 YR\b", "", base_key).strip()
        base_key = re.sub(r"\b10 YR\b", "", base_key).strip()
        base_key = re.sub(r"\s+", " ", base_key).strip()

        offer = PublicOffer(
            vendor="Global Background Screening",
            price=price,
            label=label,
            source_url=GBS_URL,
            note="Current public criminal checkout option",
        )
        for key in {base_key, label_key}:
            offers.setdefault(key, []).append(offer)
    return offers


def fetch_owens_offers() -> dict[str, list[PublicOffer]]:
    offers: dict[str, list[PublicOffer]] = {}
    headers = {"User-Agent": "Mozilla/5.0"}
    for product_type, url_template in OWENS_CRIMINAL_ENDPOINTS.items():
        page_number = 0
        while True:
            response = requests.get(url_template.format(page=page_number), headers=headers, timeout=30)
            response.raise_for_status()
            payload = response.json()
            for row in payload.get("results", []):
                price = row.get("usd")
                if price is None or price <= 0:
                    continue
                country_label = row["location"]["name"]
                key = normalize_country(country_label)
                offer = PublicOffer(
                    vendor="Owens OnLine",
                    price=float(price),
                    label=row.get("alternateName") or row.get("name") or f"Product type {product_type}",
                    source_url=url_template.format(page=page_number),
                    note=f"Current public API price, productType.id={product_type}",
                )
                offers.setdefault(key, []).append(offer)
            if not payload.get("hasNextPage"):
                break
            page_number += 1
    return offers


def pick_public_offer(
    offers_by_key: dict[str, PublicOffer] | dict[str, list[PublicOffer]],
    country: str,
    preferred_label: str | None = None,
) -> PublicOffer | None:
    keys = candidate_keys(country, preferred_label)
    for key in keys:
        offer = offers_by_key.get(key)
        if not offer:
            continue
        if isinstance(offer, list):
            return min(offer, key=lambda item: item.price)
        return offer
    return None


def currency(value: float | None) -> float | None:
    return round(value, 2) if value is not None else None


def write_headers(sheet, headers: Iterable[str]) -> None:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_table(sheet, last_column: int, last_row: int, name: str) -> None:
    reference = f"A1:{get_column_letter(last_column)}{last_row}"
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def apply_grid(sheet, start_row: int, end_row: int, end_column: int) -> None:
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_column):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def format_currency_columns(sheet, columns: Iterable[str], start_row: int, end_row: int) -> None:
    for column in columns:
        for row in range(start_row, end_row + 1):
            sheet[f"{column}{row}"].number_format = "$#,##0.00"


def format_percent_columns(sheet, columns: Iterable[str], start_row: int, end_row: int) -> None:
    for column in columns:
        for row in range(start_row, end_row + 1):
            sheet[f"{column}{row}"].number_format = "0.0%"


def set_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def build_workbook(comparison_rows: list[dict], public_source_rows: list[dict]) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True

    controls = workbook.create_sheet("Controls")
    controls["A1"] = "International Criminal Pricing Calculator"
    controls["A1"].font = Font(size=14, bold=True)
    controls["A3"] = "Compute surcharge per transaction ($)"
    controls["B3"] = DEFAULT_COMPUTE_SURCHARGE
    controls["A4"] = "Target gross margin %"
    controls["B4"] = DEFAULT_TARGET_MARGIN
    controls["A5"] = "Competitive undercut ($)"
    controls["B5"] = DEFAULT_UNDERCUT
    controls["A7"] = "Formula notes"
    controls["A8"] = "InformData cost basis = InformData offered + compute surcharge"
    controls["A9"] = "Neeyamo cost basis = Neeyamo offered + compute surcharge"
    controls["A10"] = "Target-margin sell price = cost basis / (1 - target margin)"
    controls["A11"] = "Competitive ceiling = lowest public market price - undercut"
    controls["A13"] = "Freshness"
    controls["B13"] = f"Public market prices verified on {TODAY}"
    controls["A14"] = "InformData source"
    controls["B14"] = str(INFORMDATA_SOURCE)
    controls["A15"] = "Neeyamo source"
    controls["B15"] = str(NEEYAMO_SOURCE)
    controls["A16"] = "Public sources"
    controls["B16"] = f"{GOODHIRE_URL} | {GBS_URL} | {OWENS_COUNTRIES_URL}"
    for row in range(3, 6):
        controls[f"A{row}"].fill = CONTROL_FILL
        controls[f"A{row}"].font = BOLD_FONT
        controls[f"A{row}"].border = THIN_BORDER
        controls[f"B{row}"].fill = CONTROL_FILL
        controls[f"B{row}"].border = THIN_BORDER
    controls["B3"].number_format = "$#,##0.00"
    controls["B4"].number_format = "0.0%"
    controls["B5"].number_format = "$#,##0.00"
    controls.column_dimensions["A"].width = 42
    controls.column_dimensions["B"].width = 115

    summary = workbook.create_sheet("Summary")
    summary["A1"] = "Leadership Summary"
    summary["A1"].font = Font(size=14, bold=True)
    summary["A3"] = "Scope"
    summary["B3"] = "International Criminal"
    summary["A4"] = "Default controls"
    summary["B4"] = (
        f"Compute surcharge ${DEFAULT_COMPUTE_SURCHARGE:.2f}; "
        f"target gross margin {DEFAULT_TARGET_MARGIN:.0%}; "
        f"competitive undercut ${DEFAULT_UNDERCUT:.2f}"
    )
    summary["A6"] = "Quoted countries in InformData source"
    summary["B6"] = "=COUNTA(Criminal_Comparison!A:A)-1"
    summary["A7"] = "Rows with Neeyamo criminal price"
    summary["B7"] = '=COUNT(Criminal_Comparison!E:E)'
    summary["A8"] = "Rows with a current public market price"
    summary["B8"] = '=COUNT(Criminal_Comparison!M:M)'
    summary["A9"] = "Rows where Neeyamo cost basis beats InformData cost basis"
    summary["B9"] = '=COUNTIF(Criminal_Comparison!Q:Q,"Neeyamo")'
    summary["A10"] = "Rows where InformData cost basis is at or below public market"
    summary["B10"] = '=COUNTIF(Criminal_Comparison!Y:Y,">=0")'
    summary["A11"] = "Rows where Neeyamo cost basis is at or below public market"
    summary["B11"] = '=COUNTIF(Criminal_Comparison!Z:Z,">=0")'
    summary["A13"] = "What changed from the older March 31 snapshots"
    summary["B13"] = (
        "This workbook refreshes the public benchmark layer from live GoodHire, "
        "Global Background Screening, and Owens sources on April 6, 2026. "
        "Older cached market-floor CSVs should not be used as the final benchmark."
    )
    summary["A15"] = "Top Neeyamo savings vs InformData"
    summary["A16"] = "Country"
    summary["B16"] = "Neeyamo advantage ($)"
    summary["C16"] = "InformData"
    summary["D16"] = "Neeyamo"
    summary["A24"] = "Rows where public market is below both providers"
    summary["A25"] = "Country"
    summary["B25"] = "Public floor"
    summary["C25"] = "InformData basis"
    summary["D25"] = "Neeyamo basis"

    top_neeyamo = [
        row
        for row in comparison_rows
        if row["informdata_price"] is not None and row["neeyamo_price"] is not None
    ]
    top_neeyamo.sort(key=lambda row: (row["informdata_price"] - row["neeyamo_price"]), reverse=True)
    for idx, row in enumerate(top_neeyamo[:7], start=17):
        summary[f"A{idx}"] = row["country"]
        summary[f"B{idx}"] = currency(row["informdata_price"] - row["neeyamo_price"])
        summary[f"C{idx}"] = row["informdata_price"]
        summary[f"D{idx}"] = row["neeyamo_price"]

    market_below_both = [
        row
        for row in comparison_rows
        if row["market_floor_price"] is not None
        and (
            (row["informdata_price"] is not None and row["market_floor_price"] < row["informdata_price"] + DEFAULT_COMPUTE_SURCHARGE)
            and (
                row["neeyamo_price"] is None
                or row["market_floor_price"] < row["neeyamo_price"] + DEFAULT_COMPUTE_SURCHARGE
            )
        )
    ]
    market_below_both.sort(key=lambda row: row["market_floor_price"])
    for idx, row in enumerate(market_below_both[:7], start=26):
        summary[f"A{idx}"] = row["country"]
        summary[f"B{idx}"] = row["market_floor_price"]
        summary[f"C{idx}"] = currency(row["informdata_price"] + DEFAULT_COMPUTE_SURCHARGE if row["informdata_price"] is not None else None)
        summary[f"D{idx}"] = currency(row["neeyamo_price"] + DEFAULT_COMPUTE_SURCHARGE if row["neeyamo_price"] is not None else None)

    for cell_ref in ["A3", "A4", "A6", "A7", "A8", "A9", "A10", "A11", "A13", "A15", "A24"]:
        summary[cell_ref].font = BOLD_FONT
    for row in [16, 25]:
        for column in "ABCD":
            summary[f"{column}{row}"].fill = SUBHEADER_FILL
            summary[f"{column}{row}"].font = BOLD_FONT
    format_currency_columns(summary, ["B", "C", "D"], 17, 32)
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 24
    summary.column_dimensions["C"].width = 18
    summary.column_dimensions["D"].width = 18

    comparison = workbook.create_sheet("Criminal_Comparison")
    headers = [
        "Country",
        "InformData Scope",
        "InformData Offered Price",
        "InformData Cost Basis (+ compute)",
        "Neeyamo Offered Price",
        "Neeyamo Cost Basis (+ compute)",
        "GoodHire Current Public",
        "GoodHire Label",
        "GBS Current Public",
        "GBS Label",
        "Owens Current Public",
        "Owens Label",
        "Lowest Current Public Market Price",
        "Lowest Public Vendor",
        "Lowest Public Label",
        "Lowest Internal Cost Basis",
        "Cheapest Internal Provider",
        "Price at Target Margin Using InformData",
        "Price at Target Margin Using Neeyamo",
        "Competitive Ceiling (Public Floor - Undercut)",
        "Profit at Competitive Ceiling vs InformData",
        "Margin at Competitive Ceiling vs InformData",
        "Profit at Competitive Ceiling vs Neeyamo",
        "Margin at Competitive Ceiling vs Neeyamo",
        "Gap: Competitive Ceiling - InformData Basis",
        "Gap: Competitive Ceiling - Neeyamo Basis",
        "GoodHire Source URL",
        "GBS Source URL",
        "Owens Source URL",
        "Notes",
    ]
    write_headers(comparison, headers)

    for row_number, row in enumerate(comparison_rows, start=2):
        comparison[f"A{row_number}"] = row["country"]
        comparison[f"B{row_number}"] = row["scope"]
        comparison[f"C{row_number}"] = row["informdata_price"]
        comparison[f"D{row_number}"] = f'=IF(C{row_number}="","",C{row_number}+Controls!$B$3)'
        comparison[f"E{row_number}"] = row["neeyamo_price"]
        comparison[f"F{row_number}"] = f'=IF(E{row_number}="","",E{row_number}+Controls!$B$3)'
        comparison[f"G{row_number}"] = row["goodhire_price"]
        comparison[f"H{row_number}"] = row["goodhire_label"]
        comparison[f"I{row_number}"] = row["gbs_price"]
        comparison[f"J{row_number}"] = row["gbs_label"]
        comparison[f"K{row_number}"] = row["owens_price"]
        comparison[f"L{row_number}"] = row["owens_label"]
        comparison[f"M{row_number}"] = row["market_floor_price"]
        comparison[f"N{row_number}"] = row["market_floor_vendor"]
        comparison[f"O{row_number}"] = row["market_floor_label"]
        comparison[f"P{row_number}"] = (
            f'=IF(AND(D{row_number}="",F{row_number}=""),"",'
            f'IF(D{row_number}="",F{row_number},IF(F{row_number}="",D{row_number},MIN(D{row_number},F{row_number}))))'
        )
        comparison[f"Q{row_number}"] = (
            f'=IF(P{row_number}="","",IF(AND(D{row_number}<>"",P{row_number}=D{row_number}),"InformData","Neeyamo"))'
        )
        comparison[f"R{row_number}"] = f'=IF(D{row_number}="","",D{row_number}/(1-Controls!$B$4))'
        comparison[f"S{row_number}"] = f'=IF(F{row_number}="","",F{row_number}/(1-Controls!$B$4))'
        comparison[f"T{row_number}"] = f'=IF(M{row_number}="","",M{row_number}-Controls!$B$5)'
        comparison[f"U{row_number}"] = f'=IF(OR(T{row_number}="",D{row_number}=""),"",T{row_number}-D{row_number})'
        comparison[f"V{row_number}"] = f'=IF(OR(T{row_number}="",D{row_number}=""),"",U{row_number}/T{row_number})'
        comparison[f"W{row_number}"] = f'=IF(OR(T{row_number}="",F{row_number}=""),"",T{row_number}-F{row_number})'
        comparison[f"X{row_number}"] = f'=IF(OR(T{row_number}="",F{row_number}=""),"",W{row_number}/T{row_number})'
        comparison[f"Y{row_number}"] = f'=IF(OR(T{row_number}="",D{row_number}=""),"",T{row_number}-D{row_number})'
        comparison[f"Z{row_number}"] = f'=IF(OR(T{row_number}="",F{row_number}=""),"",T{row_number}-F{row_number})'
        comparison[f"AA{row_number}"] = row["goodhire_url"]
        comparison[f"AB{row_number}"] = row["gbs_url"]
        comparison[f"AC{row_number}"] = row["owens_url"]
        comparison[f"AD{row_number}"] = row["notes"]

    last_row = len(comparison_rows) + 1
    apply_grid(comparison, 1, last_row, 30)
    comparison.freeze_panes = "A2"
    comparison.auto_filter.ref = f"A1:AD{last_row}"
    add_table(comparison, 30, last_row, "CriminalComparisonTable")
    format_currency_columns(comparison, ["C", "D", "E", "F", "G", "I", "K", "M", "P", "R", "S", "T", "U", "W", "Y", "Z"], 2, last_row)
    format_percent_columns(comparison, ["V", "X"], 2, last_row)
    set_widths(
        comparison,
        {
            "A": 24,
            "B": 14,
            "C": 18,
            "D": 22,
            "E": 18,
            "F": 22,
            "G": 18,
            "H": 22,
            "I": 18,
            "J": 28,
            "K": 18,
            "L": 28,
            "M": 22,
            "N": 18,
            "O": 30,
            "P": 20,
            "Q": 20,
            "R": 24,
            "S": 24,
            "T": 24,
            "U": 22,
            "V": 22,
            "W": 22,
            "X": 22,
            "Y": 22,
            "Z": 22,
            "AA": 26,
            "AB": 26,
            "AC": 26,
            "AD": 40,
        },
    )
    for row in range(2, last_row + 1):
        comparison[f"AD{row}"].alignment = Alignment(vertical="top", wrap_text=True)

    source_sheet = workbook.create_sheet("Current_Public_Sources")
    source_headers = ["Vendor", "Country Key", "Label", "Price", "Source URL", "Note"]
    write_headers(source_sheet, source_headers)
    for row_number, row in enumerate(public_source_rows, start=2):
        source_sheet[f"A{row_number}"] = row["vendor"]
        source_sheet[f"B{row_number}"] = row["country_key"]
        source_sheet[f"C{row_number}"] = row["label"]
        source_sheet[f"D{row_number}"] = row["price"]
        source_sheet[f"E{row_number}"] = row["source_url"]
        source_sheet[f"F{row_number}"] = row["note"]
    source_last_row = len(public_source_rows) + 1
    apply_grid(source_sheet, 1, source_last_row, 6)
    source_sheet.freeze_panes = "A2"
    add_table(source_sheet, 6, source_last_row, "PublicSourceTable")
    format_currency_columns(source_sheet, ["D"], 2, source_last_row)
    set_widths(
        source_sheet,
        {"A": 28, "B": 24, "C": 42, "D": 14, "E": 42, "F": 32},
    )

    return workbook


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    informdata_rows = load_informdata_rows()
    old_owens_matches = load_old_owens_matches()
    neeyamo_prices = load_neeyamo_prices()
    goodhire_offers = fetch_goodhire_offers()
    gbs_offers = fetch_gbs_offers()
    owens_offers = fetch_owens_offers()

    public_source_rows: list[dict] = []
    for country_key, offer in goodhire_offers.items():
        public_source_rows.append(
            {
                "vendor": offer.vendor,
                "country_key": country_key,
                "label": offer.label,
                "price": offer.price,
                "source_url": offer.source_url,
                "note": offer.note,
            }
        )
    for country_key, offers in gbs_offers.items():
        for offer in offers:
            public_source_rows.append(
                {
                    "vendor": offer.vendor,
                    "country_key": country_key,
                    "label": offer.label,
                    "price": offer.price,
                    "source_url": offer.source_url,
                    "note": offer.note,
                }
            )
    for country_key, offers in owens_offers.items():
        for offer in offers:
            public_source_rows.append(
                {
                    "vendor": offer.vendor,
                    "country_key": country_key,
                    "label": offer.label,
                    "price": offer.price,
                    "source_url": offer.source_url,
                    "note": offer.note,
                }
            )

    comparison_rows: list[dict] = []
    for row in informdata_rows:
        country = row["country"]
        scope = row["scope"]
        informdata_price = row["informdata_price"]

        neeyamo_entry = None
        for key in candidate_keys(country):
            if key in neeyamo_prices:
                neeyamo_entry = neeyamo_prices[key]
                break

        goodhire_offer = pick_public_offer(goodhire_offers, country, row["old_goodhire_label"])
        gbs_offer = pick_public_offer(gbs_offers, country, row["old_gbs_label"])
        owens_offer = pick_public_offer(
            owens_offers,
            country,
            old_owens_matches.get((country, ""), ""),
        )

        public_candidates = [offer for offer in [goodhire_offer, gbs_offer, owens_offer] if offer]
        market_floor = min(public_candidates, key=lambda item: item.price) if public_candidates else None

        notes: list[str] = []
        if not neeyamo_entry:
            notes.append("No Neeyamo criminal rate matched from the current local rate card.")
        if not market_floor:
            notes.append("No current public market benchmark matched from GoodHire / GBS / Owens.")
        elif informdata_price is not None and market_floor.price < informdata_price + DEFAULT_COMPUTE_SURCHARGE:
            notes.append("Current public market ceiling is below the default InformData cost basis.")
        if market_floor and neeyamo_entry and market_floor.price < neeyamo_entry["price"] + DEFAULT_COMPUTE_SURCHARGE:
            notes.append("Current public market ceiling is below the default Neeyamo cost basis.")

        comparison_rows.append(
            {
                "country": country,
                "scope": scope,
                "informdata_price": currency(informdata_price),
                "neeyamo_price": currency(neeyamo_entry["price"]) if neeyamo_entry else None,
                "goodhire_price": currency(goodhire_offer.price) if goodhire_offer else None,
                "goodhire_label": goodhire_offer.label if goodhire_offer else "",
                "goodhire_url": goodhire_offer.source_url if goodhire_offer else "",
                "gbs_price": currency(gbs_offer.price) if gbs_offer else None,
                "gbs_label": gbs_offer.label if gbs_offer else "",
                "gbs_url": gbs_offer.source_url if gbs_offer else "",
                "owens_price": currency(owens_offer.price) if owens_offer else None,
                "owens_label": owens_offer.label if owens_offer else "",
                "owens_url": owens_offer.source_url if owens_offer else "",
                "market_floor_price": currency(market_floor.price) if market_floor else None,
                "market_floor_vendor": market_floor.vendor if market_floor else "",
                "market_floor_label": market_floor.label if market_floor else "",
                "notes": " ".join(notes),
            }
        )

    workbook = build_workbook(comparison_rows, public_source_rows)
    workbook_path = OUTPUT_DIR / f"International_Criminal_Pricing_Calculator_{TODAY}.xlsx"
    workbook.save(workbook_path)

    csv_path = OUTPUT_DIR / f"International_Criminal_Pricing_Calculator_{TODAY}.csv"
    csv_fields = [
        "country",
        "scope",
        "informdata_price",
        "neeyamo_price",
        "goodhire_price",
        "goodhire_label",
        "gbs_price",
        "gbs_label",
        "owens_price",
        "owens_label",
        "market_floor_price",
        "market_floor_vendor",
        "market_floor_label",
        "notes",
    ]
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=csv_fields,
        )
        writer.writeheader()
        writer.writerows([{field: row.get(field) for field in csv_fields} for row in comparison_rows])

    neeyamo_overlap = sum(1 for row in comparison_rows if row["neeyamo_price"] is not None)
    market_overlap = sum(1 for row in comparison_rows if row["market_floor_price"] is not None)
    neeyamo_better = sum(
        1
        for row in comparison_rows
        if row["informdata_price"] is not None
        and row["neeyamo_price"] is not None
        and row["neeyamo_price"] < row["informdata_price"]
    )
    market_below_informdata = sum(
        1
        for row in comparison_rows
        if row["informdata_price"] is not None
        and row["market_floor_price"] is not None
        and row["market_floor_price"] < row["informdata_price"] + DEFAULT_COMPUTE_SURCHARGE
    )

    summary = {
        "workbook": str(workbook_path),
        "csv": str(csv_path),
        "quoted_rows": len(comparison_rows),
        "neeyamo_overlap": neeyamo_overlap,
        "market_overlap": market_overlap,
        "neeyamo_cheaper_than_informdata": neeyamo_better,
        "public_market_below_informdata_basis": market_below_informdata,
    }
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
